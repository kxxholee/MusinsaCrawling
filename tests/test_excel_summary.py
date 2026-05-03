from __future__ import annotations

import unittest

from openpyxl import Workbook

from musinsa.excel import build_sku_summary_sheet, split_report_model_name_and_color
from musinsa.schemas import ProductRow, SkuRow


def product(product_id: str, name: str) -> ProductRow:
    return ProductRow(
        crawl_date="2026-05-03",
        gender_filter="F",
        brand="musinsastandardwoman",
        major_category="상의",
        raw_category="반소매 티셔츠",
        category_code="001",
        product_id=product_id,
        product_name=name,
        product_url=f"https://www.musinsa.com/products/{product_id}?gf=F",
        price=19900,
        discount_rate=None,
        image_url="",
        is_soldout="",
        source_category_url="",
    )


def sku(product_id: str, name: str, color: str, size: str) -> SkuRow:
    return SkuRow(
        major_category="상의",
        raw_category="반소매 티셔츠",
        product_id=product_id,
        product_name=name,
        color=color,
        size=size,
        sku_key=f"{product_id}_{color}_{size}",
        sku_status="판매중",
        product_url=f"https://www.musinsa.com/products/{product_id}?gf=F",
    )


class SkuSummarySheetTest(unittest.TestCase):
    def test_split_report_model_name_and_color_uses_rightmost_suffix(self) -> None:
        name, color = split_report_model_name_and_color("우먼즈 티셔츠 [기획] [블랙]")

        self.assertEqual(name, "우먼즈 티셔츠 [기획]")
        self.assertEqual(color, "블랙")

    def test_summary_sheet_splits_colors_and_adds_model_sku_count(self) -> None:
        names = {
            "1001": "무신사 스탠다드 (MUSINSA STANDARD) 우먼즈 베이식 티셔츠 [화이트]",
            "1002": "무신사 스탠다드 (MUSINSA STANDARD) 우먼즈 베이식 티셔츠 [블랙]",
            "1003": "무신사 스탠다드 (MUSINSA STANDARD) 우먼즈 베이식 티셔츠 [크림]",
            "1004": "무신사 스탠다드 (MUSINSA STANDARD) 우먼즈 베이식 티셔츠 [그레이]",
        }
        products = [product(product_id, name) for product_id, name in names.items()]
        sku_rows = [
            sku(product_id, name, color, size)
            for product_id, name, color in [
                ("1001", names["1001"], "화이트"),
                ("1002", names["1002"], "블랙"),
                ("1003", names["1003"], "크림"),
            ]
            for size in ("S", "M", "L")
        ]

        wb = Workbook()
        ws = wb.active
        build_sku_summary_sheet(
            ws,
            products,
            sku_rows,
            category_tree=[
                {
                    "major_name": "상의",
                    "major_code": "001",
                    "subs": [
                        {"code": "001001", "name": "반소매 티셔츠"},
                        {"code": "001002", "name": "스커트"},
                    ],
                }
            ],
        )

        self.assertEqual(
            [ws.cell(1, col).value for col in range(1, 12)],
            [
                "카테고리", "중분류", "모델명", "색상", "컬러 수", "사이즈 수",
                "모델 SKU 개수", "SKU 개수", "SKU 비중", "최고/최저/평균가",
                "베이직/뉴베이직/트렌디",
            ],
        )
        self.assertEqual(ws["C2"].value, "우먼즈 베이식 티셔츠")
        self.assertEqual([ws.cell(row, 4).value for row in range(2, 5)], ["화이트", "블랙", "크림"])
        self.assertEqual(ws["E2"].value, 3)
        self.assertEqual(ws["F2"].value, 3)
        self.assertEqual(ws["G2"].value, 9)
        self.assertEqual(ws["H2"].value, 9)
        self.assertEqual(ws["I2"].number_format, "0.00%")
        self.assertEqual(ws.max_row, 4)
        self.assertNotIn("그레이", [ws.cell(row, 4).value for row in range(2, ws.max_row + 1)])
        self.assertNotIn("스커트", [ws.cell(row, 2).value for row in range(2, ws.max_row + 1)])

        merged_ranges = {str(cell_range) for cell_range in ws.merged_cells.ranges}
        self.assertTrue({"C2:C4", "E2:E4", "F2:F4", "G2:G4"}.issubset(merged_ranges))
        self.assertTrue({"B2:B4", "H2:H4", "I2:I4", "J2:J4", "K2:K4"}.issubset(merged_ranges))

    def test_summary_sheet_skips_models_with_zero_model_sku_count(self) -> None:
        products = [
            product("2001", "무신사 스탠다드 (MUSINSA STANDARD) 우먼즈 단종 티셔츠 [블랙]"),
        ]

        wb = Workbook()
        ws = wb.active
        build_sku_summary_sheet(
            ws,
            products,
            [],
            category_tree=[
                {
                    "major_name": "상의",
                    "major_code": "001",
                    "subs": [{"code": "001001", "name": "반소매 티셔츠"}],
                }
            ],
        )

        self.assertEqual(ws.max_row, 1)
        self.assertEqual(ws.max_column, 11)


if __name__ == "__main__":
    unittest.main()
