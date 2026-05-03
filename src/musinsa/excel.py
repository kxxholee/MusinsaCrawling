from __future__ import annotations

import re
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Set, Tuple

import pandas as pd

from .categories import category_tree_to_master_rows
from .schemas import OptionRow, ProductRow, SkuRow
from .utils import clean_text, dataclass_list_to_dicts, ensure_gender_filter_url


COLOR_SUFFIX_RE = re.compile(r"\[([^\[\]]+)\]\s*$")


def make_model_key_from_values(
    major_category: Any,
    raw_category: Any,
    product_id: Any,
    product_url: Any,
) -> str:
    major_value = "" if pd.isna(major_category) else major_category
    raw_value = "" if pd.isna(raw_category) else raw_category
    product_id_value = "" if pd.isna(product_id) else product_id
    product_url_value = "" if pd.isna(product_url) else product_url
    product_identity = clean_text(product_id_value) or ensure_gender_filter_url(product_url_value)
    return "|".join(
        [
            clean_text(major_value),
            clean_text(raw_value),
            product_identity,
        ]
    )


def add_model_key_column(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        df["_model_key"] = []
        return df

    df = df.copy()
    df["_model_key"] = df.apply(
        lambda row: make_model_key_from_values(
            row.get("major_category", ""),
            row.get("raw_category", ""),
            row.get("product_id", ""),
            row.get("product_url", ""),
        ),
        axis=1,
    )
    return df


def normalize_url_columns(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    for column in ("product_url", "source_category_url"):
        if column in df.columns:
            df[column] = df[column].map(
                lambda value: "" if pd.isna(value) else ensure_gender_filter_url(value)
            )
    return df


def clean_model_name(raw_name: str) -> str:
    """보고서 표시용 모델명에서 브랜드 prefix와 ' - 사이즈 & 후기' 등을 제거."""
    if not raw_name:
        return ""
    name = clean_text(raw_name)
    name = re.sub(r"^무신사\s*스탠다드\s*\(MUSINSA\s*STANDARD\)\s*", "", name, flags=re.IGNORECASE)
    name = re.sub(r"\s*-\s*사이즈\s*&?\s*후기.*$", "", name)
    name = re.sub(r"\s*-\s*후기.*$", "", name)
    return name.strip()


def split_report_model_name_and_color(raw_name: str) -> Tuple[str, str]:
    """상품명 끝의 [색상] suffix 를 SKU_요약용 모델명/색상으로 분리."""
    if not raw_name:
        return "", ""

    name = clean_text(raw_name)
    match = COLOR_SUFFIX_RE.search(name)
    if not match:
        return clean_model_name(name), ""

    model_name = clean_model_name(name[:match.start()].strip())
    color = clean_text(match.group(1))
    return model_name, color


def format_price_summary(prices: Iterable[Optional[int]]) -> str:
    """중분류 단위 가격을 '최고/최저/평균원' 형식으로."""
    valid = [int(p) for p in prices if p is not None]
    if not valid:
        return "-"
    return f"{max(valid):,} / {min(valid):,} / {int(round(sum(valid) / len(valid))):,}원"


def build_model_summary_df(
    products: List[ProductRow],
    sku_rows: List[SkuRow],
) -> pd.DataFrame:
    """모델(상품) 단위 요약 — SKU_요약 시트와 model_summary 시트 공통 입력."""
    columns = [
        "major_category", "raw_category", "product_id", "product_name",
        "color_count", "size_count", "sku_count",
        "available_sku_count", "soldout_sku_count",
        "price", "product_url",
    ]

    if not products:
        return pd.DataFrame(columns=columns)

    prod_df = normalize_url_columns(pd.DataFrame(dataclass_list_to_dicts(products)))
    sku_df = normalize_url_columns(pd.DataFrame(dataclass_list_to_dicts(sku_rows)))
    prod_df = add_model_key_column(prod_df)

    if sku_df.empty:
        prod_df = prod_df.assign(
            color_count=0, size_count=0, sku_count=0,
            available_sku_count=0, soldout_sku_count=0,
        )
    else:
        sku_df = add_model_key_column(sku_df)
        agg = sku_df.groupby("_model_key").agg(
            color_count=("color", "nunique"),
            size_count=("size", "nunique"),
            sku_count=("_model_key", "count"),
            available_sku_count=("sku_status", lambda s: int((s == "판매중").sum())),
            soldout_sku_count=("sku_status", lambda s: int((s == "품절").sum())),
        ).reset_index()
        prod_df = prod_df.merge(agg, on="_model_key", how="left")
        for col in ("color_count", "size_count", "sku_count", "available_sku_count", "soldout_sku_count"):
            prod_df[col] = prod_df[col].fillna(0).astype(int)

    return prod_df.reindex(columns=columns)


def build_sku_summary_sheet(
    ws: Any,
    products: List[ProductRow],
    sku_rows: List[SkuRow],
    category_tree: Optional[List[Dict[str, Any]]] = None,
) -> None:
    """style.xlsx 양식의 SKU_요약 시트 — 모델명 색상 suffix 를 별도 열로 분리."""
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    headers = [
        "카테고리", "중분류", "모델명", "색상", "컬러 수", "사이즈 수",
        "모델 SKU 개수", "SKU 개수", "SKU 비중", "최고/최저/평균가",
        "베이직/뉴베이직/트렌디",
    ]
    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="D9D9D9")
    header_font = Font(bold=True, size=11)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)

    for col_idx, label in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.alignment = center
        cell.fill = header_fill
        cell.border = border
    ws.row_dimensions[1].height = 30
    widths = {
        1: 18, 2: 16, 3: 38, 4: 16, 5: 8, 6: 8,
        7: 14, 8: 12, 9: 12, 10: 26, 11: 22,
    }
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"

    summary_df = build_model_summary_df(products, sku_rows)
    if summary_df.empty:
        return

    summary_df = add_model_key_column(summary_df)
    report_parts = summary_df["product_name"].map(split_report_model_name_and_color)
    summary_df["_report_model_name"] = report_parts.map(lambda part: part[0])
    summary_df["_report_color"] = report_parts.map(lambda part: part[1])
    summary_df["_report_group_key"] = summary_df.apply(
        lambda row: "|".join(
            [
                clean_text(row.get("major_category", "")),
                clean_text(row.get("raw_category", "")),
                clean_text(row.get("_report_model_name", "")),
            ]
        ),
        axis=1,
    )
    summary_df = summary_df[summary_df["sku_count"] > 0]
    if summary_df.empty:
        return

    size_count_by_report_key: Dict[str, int] = {}
    if sku_rows:
        sku_df = add_model_key_column(
            normalize_url_columns(pd.DataFrame(dataclass_list_to_dicts(sku_rows)))
        )
        sku_df = sku_df.merge(
            summary_df[["_model_key", "_report_group_key"]],
            on="_model_key",
            how="left",
        )
        size_count_by_report_key = (
            sku_df.dropna(subset=["_report_group_key"])
            .groupby("_report_group_key")["size"]
            .nunique()
            .astype(int)
            .to_dict()
        )

    color_count_by_report_key = (
        summary_df.groupby("_report_group_key")["_report_color"]
        .nunique()
        .astype(int)
        .to_dict()
    )
    model_sku_count_by_report_key: Dict[str, int] = {
        report_key: color_count * size_count_by_report_key.get(report_key, 0)
        for report_key, color_count in color_count_by_report_key.items()
    }
    summary_df = summary_df[
        summary_df["_report_group_key"].map(
            lambda report_key: model_sku_count_by_report_key.get(str(report_key), 0) > 0
        )
    ]
    if summary_df.empty:
        return

    total_sku = int(summary_df["sku_count"].sum())

    tree = category_tree or []
    major_order: List[str] = [m["major_name"] for m in tree]
    raw_order_by_major: Dict[str, List[str]] = {
        m["major_name"]: [s["name"] for s in m["subs"]] for m in tree
    }

    for extra in summary_df["major_category"].dropna().unique().tolist():
        if extra and extra not in major_order:
            major_order.append(extra)
            raw_order_by_major.setdefault(extra, [])

    current_row = 2
    for major_cat in major_order:
        major_df = summary_df[summary_df["major_category"] == major_cat]

        major_total = int(major_df["sku_count"].sum())
        major_pct = major_total / total_sku if total_sku else 0

        raw_master = raw_order_by_major.get(major_cat, [])
        raw_in_data = list(major_df["raw_category"].unique())
        raw_sequence = (
            raw_master
            + [r for r in raw_in_data if r not in raw_master]
        )
        if not raw_sequence:
            continue

        major_start = current_row

        for raw_cat in raw_sequence:
            raw_df = major_df[major_df["raw_category"] == raw_cat].sort_values("product_id")

            raw_total = int(raw_df["sku_count"].sum())
            raw_pct = raw_total / total_sku if total_sku else 0
            price_summary = format_price_summary(raw_df["price"].tolist())
            style_summary = "미분류: 100%"

            raw_start = current_row
            if raw_df.empty:
                ws.cell(row=current_row, column=3, value="")
                ws.cell(row=current_row, column=4, value="")
                ws.cell(row=current_row, column=5, value=0)
                ws.cell(row=current_row, column=6, value=0)
                ws.cell(row=current_row, column=7, value=0)
                current_row += 1
            else:
                for _, model_df in raw_df.groupby("_report_group_key", sort=False):
                    model_start = current_row
                    model_name = str(model_df["_report_model_name"].iloc[0])
                    colors = model_df["_report_color"].drop_duplicates().tolist()
                    color_count = len(colors)
                    report_key = str(model_df["_report_group_key"].iloc[0])
                    size_count = size_count_by_report_key.get(report_key, 0)
                    model_sku_count = model_sku_count_by_report_key.get(report_key, 0)

                    for color in colors:
                        ws.cell(row=current_row, column=4, value=color)
                        current_row += 1

                    model_end = current_row - 1
                    ws.cell(row=model_start, column=3, value=model_name)
                    ws.cell(row=model_start, column=5, value=color_count)
                    ws.cell(row=model_start, column=6, value=int(size_count))
                    ws.cell(row=model_start, column=7, value=model_sku_count)

                    if model_end > model_start:
                        for col in (3, 5, 6, 7):
                            ws.merge_cells(
                                start_row=model_start, end_row=model_end,
                                start_column=col, end_column=col,
                            )
            raw_end = current_row - 1

            ws.cell(row=raw_start, column=2, value=raw_cat)
            ws.cell(row=raw_start, column=8, value=raw_total)
            pct_cell = ws.cell(row=raw_start, column=9, value=raw_pct)
            pct_cell.number_format = "0.00%"
            ws.cell(row=raw_start, column=10, value=price_summary)
            ws.cell(row=raw_start, column=11, value=style_summary)

            if raw_end > raw_start:
                for col in (2, 8, 9, 10, 11):
                    ws.merge_cells(
                        start_row=raw_start, end_row=raw_end,
                        start_column=col, end_column=col,
                    )

        major_end = current_row - 1
        if major_end >= major_start:
            ws.cell(
                row=major_start,
                column=1,
                value=f"{major_cat}\n(총 SKU: {major_total:,} / {major_pct:.1%})",
            )
            if major_end > major_start:
                ws.merge_cells(
                    start_row=major_start, end_row=major_end,
                    start_column=1, end_column=1,
                )

    last_row = current_row - 1
    if last_row >= 2:
        for row in ws.iter_rows(min_row=2, max_row=last_row, min_col=1, max_col=11):
            for cell in row:
                cell.border = border
                cell.alignment = left_wrap if cell.column == 3 else center


_URL_COLUMNS = ("product_url", "source_category_url")


def _normalize_cell_value(column: str, value: Any) -> Any:
    """openpyxl 셀에 그대로 넣을 수 있게 None/URL 정규화."""
    if value is None:
        return ""
    if column in _URL_COLUMNS and isinstance(value, str) and value:
        return ensure_gender_filter_url(value)
    return value


def _append_dataclass_sheet(ws: Any, rows: Iterable[Any], columns: List[str]) -> None:
    """ProductRow/OptionRow/SkuRow 같은 dataclass 리스트를 헤더 + 데이터로 append."""
    ws.append(columns)
    for row in rows:
        record = row.__dict__
        ws.append([_normalize_cell_value(col, record.get(col)) for col in columns])


def _append_dict_sheet(ws: Any, rows: Iterable[Dict[str, Any]], columns: List[str]) -> None:
    ws.append(columns)
    for row in rows:
        ws.append([_normalize_cell_value(col, row.get(col)) for col in columns])


def _append_model_summary_sheet(ws: Any, df: "pd.DataFrame") -> None:
    columns = list(df.columns)
    ws.append(columns)

    if df.empty:
        return

    for record in df.to_dict(orient="records"):
        ws.append([_normalize_cell_value(col, record.get(col)) for col in columns])


def _apply_raw_sheet_styles(ws: Any) -> None:
    """헤더 스타일 / 첫 행 고정 / 자동 필터 / 자동 컬럼 폭 적용 (워크북 인메모리)."""
    from openpyxl.styles import Alignment, Font, PatternFill

    if ws.max_row < 1 or ws.max_column < 1:
        return

    header_fill = PatternFill("solid", fgColor="EDEDED")
    header_font = Font(bold=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for column_cells in ws.columns:
        max_len = 0
        col_letter = column_cells[0].column_letter
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            if len(value) > max_len:
                max_len = len(value)
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 45)


def adjust_excel_format(output_path: Path, skip_sheets: Optional[Set[str]] = None) -> None:
    """기존 파일에 raw 시트 스타일을 다시 적용 (외부 진입점). SKU_요약은 건드리지 않음."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        return

    skip = skip_sheets or set()
    wb = load_workbook(output_path)

    for ws in wb.worksheets:
        if ws.title in skip:
            continue
        _apply_raw_sheet_styles(ws)

    wb.save(output_path)


def save_excel(
    output_path: Path,
    products: List[ProductRow],
    options: List[OptionRow],
    sku_rows: List[SkuRow],
    category_tree: Optional[List[Dict[str, Any]]] = None,
) -> None:
    """SKU_요약(보고서형)을 맨 앞에, 모델/원본 시트를 뒤에 두는 단일 패스 워크북.

    pandas DataFrame 사본 + 두 번째 load_workbook 패스를 모두 없애 피크 메모리를
    크게 줄였다. SKU_요약 시트는 random-access cell 쓰기가 필요해 일반 모드를
    유지하지만, raw 시트들은 dataclass → tuple → ws.append() 로 스트리밍한다.
    """
    from openpyxl import Workbook

    output_path.parent.mkdir(parents=True, exist_ok=True)

    # 모델 단위 집계는 pandas groupby 가 단순해서 여기서만 DataFrame 사용.
    model_summary_df = build_model_summary_df(products, sku_rows)
    if not model_summary_df.empty:
        model_summary_df = normalize_url_columns(model_summary_df)

    wb = Workbook()
    # 기본 생성 시트는 SKU_요약 이름으로 곧바로 사용 (시트 인덱스 0).
    summary_ws = wb.active
    summary_ws.title = "SKU_요약"
    build_sku_summary_sheet(summary_ws, products, sku_rows, category_tree=category_tree)

    model_ws = wb.create_sheet("model_summary")
    _append_model_summary_sheet(model_ws, model_summary_df)
    _apply_raw_sheet_styles(model_ws)

    products_ws = wb.create_sheet("products_raw")
    _append_dataclass_sheet(
        products_ws, products, list(ProductRow.__annotations__.keys())
    )
    _apply_raw_sheet_styles(products_ws)

    options_ws = wb.create_sheet("options_raw")
    _append_dataclass_sheet(
        options_ws, options, list(OptionRow.__annotations__.keys())
    )
    _apply_raw_sheet_styles(options_ws)

    sku_ws = wb.create_sheet("sku_result")
    _append_dataclass_sheet(sku_ws, sku_rows, list(SkuRow.__annotations__.keys()))
    _apply_raw_sheet_styles(sku_ws)

    category_ws = wb.create_sheet("category_master")
    _append_dict_sheet(
        category_ws,
        category_tree_to_master_rows(category_tree or []),
        ["major_category", "major_code", "category_code", "raw_category"],
    )
    _apply_raw_sheet_styles(category_ws)

    wb.save(output_path)
