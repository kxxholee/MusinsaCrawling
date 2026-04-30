#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
musinsa_sku_generator_demo.py

무신사 스탠다드 여성 상품 SKU 계산용 엑셀 생성기 데모.

설치:
    uv sync

테스트 실행:
    uv run python musinsa_sku_generator_demo.py --max-products 10 --headless

전체 실행 예시:
    uv run python musinsa_sku_generator_demo.py --max-products 0 --headless

주의:
    - 이 파일은 "데모"입니다.
    - 무신사 화면 구조가 바뀌면 상품명, 가격, 옵션 선택자 부분을 수정해야 할 수 있습니다.
    - 여성 필터 gf=F는 모든 URL에 항상 붙도록 고정되어 있습니다.
    - 너무 빠른 요청은 사이트에 부담을 줄 수 있으므로 delay 값을 너무 낮추지 마세요.
"""

from __future__ import annotations

import argparse
import re
import shutil
import sys
import time
from dataclasses import dataclass
from datetime import datetime, timezone, timedelta
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Set, Tuple
from urllib.parse import parse_qsl, urlencode, urlsplit, urlunsplit


# =========================
# 1. 외부 라이브러리 확인
# =========================

try:
    import pandas as pd
except ImportError:
    print("[오류] pandas가 설치되어 있지 않습니다.")
    print("설치 명령: pip install pandas openpyxl tqdm selenium beautifulsoup4")
    sys.exit(1)

try:
    from tqdm import tqdm
except ImportError:
    print("[오류] tqdm이 설치되어 있지 않습니다.")
    print("설치 명령: pip install pandas openpyxl tqdm selenium beautifulsoup4")
    sys.exit(1)

try:
    from bs4 import BeautifulSoup
except ImportError:
    print("[오류] beautifulsoup4가 설치되어 있지 않습니다.")
    print("설치 명령: pip install beautifulsoup4")
    sys.exit(1)

try:
    from selenium import webdriver
    from selenium.common.exceptions import TimeoutException, WebDriverException
    from selenium.webdriver.common.by import By
    from selenium.webdriver.firefox.options import Options as FirefoxOptions
    from selenium.webdriver.firefox.service import Service as FirefoxService
    from selenium.webdriver.support.ui import WebDriverWait
except ImportError:
    print("[오류] selenium이 설치되어 있지 않습니다.")
    print("설치 명령: pip install selenium beautifulsoup4")
    sys.exit(1)


# =========================
# 2. 고정 설정값
# =========================

BRAND = "musinsastandardwoman"
GENDER_FILTER = "F"  # 여성 필터 고정. 모든 URL에 반드시 들어감.
BASE_BRAND_URL = f"https://www.musinsa.com/brand/{BRAND}/products"

DEFAULT_OUTPUT = "musinsa_standard_female_sku.xlsx"

KST = timezone(timedelta(hours=9))


# 무신사 표준 대분류 코드 → 한글명 폴백.
# 디스커버리에서 a 태그 텍스트를 못 가져왔을 때만 사용합니다.
MAJOR_NAME_FALLBACK: Dict[str, str] = {
    "001": "상의",
    "002": "아우터",
    "003": "바지",
    "020": "스커트",
    "022": "원피스",
    "100": "원피스/스커트",
    "101": "소품",
    "102": "모자",
    "103": "신발",
    "104": "가방",
    "005": "스포츠/레저",
}

# 디스커버리가 랜딩 페이지에서 6자리 서브 코드를 못 찾았을 때
# 직접 방문해서 서브 네비게이션을 긁을 대분류 후보 시드.
MAJOR_CODE_SEEDS: List[str] = [
    "001", "002", "003", "020", "022", "100", "101", "102", "103", "104", "005",
]

# 색상 후보. 키는 표준 색상명, 값은 같은 색을 가리키는 한·영 표기 목록입니다.
# 옵션 텍스트에서 어떤 alias가 매치되든 표준 키로 정규화되어 같은 색상으로 묶입니다.
COLOR_ALIASES: Dict[str, List[str]] = {
    "블랙": ["BLACK", "블랙"],
    "화이트": ["WHITE", "화이트"],
    "아이보리": ["IVORY", "아이보리"],
    "그레이": ["GRAY", "GREY", "그레이"],
    "멜란지": ["멜란지"],
    "네이비": ["NAVY", "네이비"],
    "스카이블루": ["스카이블루"],
    "라이트블루": ["라이트블루"],
    "블루": ["BLUE", "블루"],
    "베이지": ["BEIGE", "베이지"],
    "브라운": ["BROWN", "브라운"],
    "카키": ["KHAKI", "카키"],
    "그린": ["GREEN", "그린"],
    "핑크": ["PINK", "핑크"],
    "레드": ["RED", "레드"],
    "옐로우": ["YELLOW", "옐로우"],
    "퍼플": ["PURPLE", "퍼플"],
    "오렌지": ["ORANGE", "오렌지"],
    "차콜": ["CHARCOAL", "차콜"],
    "크림": ["CREAM", "크림"],
    "오트밀": ["OATMEAL", "오트밀"],
    "민트": ["MINT", "민트"],
    "데님": ["DENIM", "데님"],
}

SIZE_WORD_RE = re.compile(
    r"(?<![A-Z0-9])("
    r"XXS|XS|S|M|L|XL|XXL|XXXL|"
    r"FREE|ONE\s*SIZE|"
    r"W\d{2}|"
    r"\d{2,3}"
    r")(?![A-Z0-9])",
    re.IGNORECASE,
)

# 의류 사이즈로 의미 있는 숫자만 통과시킵니다. "리뷰 106만 개" 같은 본문 노이즈를 차단합니다.
NUMERIC_SIZE_WHITELIST = {
    # 인치 (바지 허리)
    "24", "25", "26", "27", "28", "29", "30", "31", "32", "33", "34", "35", "36", "37", "38", "40", "42", "44",
    # 한국식 상의 사이즈
    "85", "90", "95", "100", "105", "110", "115",
}

# 사이즈 후보 텍스트에 이 단어들이 섞여 있으면 가격/리뷰/쿠폰 노이즈로 보고 거부합니다.
SIZE_NOISE_WORDS = ("원", "%", "만", "개", "리뷰", "쿠폰", "포인트", "회", "건", "별점", "평점")

# 실제 구매 옵션이 아닌 Q&A/리뷰/착용 예시 텍스트를 SKU 후보에서 제외합니다.
OPTION_NOISE_WORDS = (
    "상품상세문의", "답변완료", "리뷰", "후기", "착용", "문의", "별점", "평점",
    "좋아요", "신고", "작성자", "사이즈 추천", "구매 후기",
    "상품명", "체형정보", "cm", "kg", "신체정보",
)
OPTION_DATE_RE = re.compile(r"\b(?:\d{2,4}[./-])?\d{1,2}[./-]\d{1,2}\b")

# "원"을 필수로 두어 누적 리뷰수/평점 같은 본문 숫자가 가격으로 잡히지 않게 합니다.
PRICE_RE = re.compile(r"(?<![0-9])([0-9]{1,3}(?:,[0-9]{3})+|[0-9]{4,7})\s*원")

# 무신사 의류 가격의 합리적 범위. 이 바깥은 가격 후보에서 제외합니다.
PRICE_MIN = 1_000
PRICE_MAX = 5_000_000
PRODUCT_ID_PATTERNS = [
    re.compile(r"/products/(\d+)"),
    re.compile(r"/goods/(\d+)"),
    re.compile(r"goods_no=(\d+)"),
    re.compile(r"goodsNo=(\d+)"),
]


# =========================
# 3. 자료 구조
# =========================

@dataclass
class ProductRow:
    crawl_date: str
    gender_filter: str
    brand: str
    major_category: str
    raw_category: str
    category_code: str
    product_id: str
    product_name: str
    product_url: str
    price: Optional[int]
    discount_rate: Optional[str]
    image_url: str
    is_soldout: str
    source_category_url: str


@dataclass
class OptionRow:
    product_id: str
    product_name: str
    major_category: str
    raw_category: str
    color: str
    size: str
    option_name: str
    option_status: str
    option_price: Optional[int]
    stock_qty: Optional[int]
    product_url: str


@dataclass
class SkuRow:
    major_category: str
    raw_category: str
    product_id: str
    product_name: str
    color: str
    size: str
    sku_key: str
    sku_status: str
    product_url: str


# =========================
# 4. 공통 유틸
# =========================

def now_kst_str() -> str:
    return datetime.now(KST).strftime("%Y-%m-%d %H:%M:%S")


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def parse_int_price(text: str) -> Optional[int]:
    """
    텍스트 안의 모든 가격 후보 중 합리적 범위에 들어오는 값의 최댓값을 반환합니다.
    카드 상단의 쿠폰가/할인가가 본가보다 먼저 노출되는 경우가 많아,
    "처음 매치"가 아니라 "가장 큰 값"을 본가 후보로 채택합니다.
    """
    text = clean_text(text)
    candidates: List[int] = []
    for match in PRICE_RE.finditer(text):
        raw = match.group(1).replace(",", "")
        try:
            value = int(raw)
        except ValueError:
            continue
        if PRICE_MIN <= value <= PRICE_MAX:
            candidates.append(value)
    if not candidates:
        return None
    return max(candidates)


def parse_discount(text: str) -> Optional[str]:
    text = clean_text(text)
    match = re.search(r"(\d{1,3})\s*%", text)
    if not match:
        return None
    return f"{match.group(1)}%"


def extract_product_id(url: str) -> str:
    for pattern in PRODUCT_ID_PATTERNS:
        match = pattern.search(url or "")
        if match:
            return match.group(1)
    return ""


def ensure_gender_filter_url(url: str) -> str:
    """
    무신사 URL에는 여성 필터 gf=F를 반드시 포함합니다.
    기존 query string은 유지하되 gf 값만 고정값으로 교체합니다.
    """
    url = clean_text(url)
    if not url:
        return ""

    parts = urlsplit(url)
    query_pairs = [
        (key, value)
        for key, value in parse_qsl(parts.query, keep_blank_values=True)
        if key != "gf"
    ]
    query_pairs.append(("gf", GENDER_FILTER))
    return urlunsplit(
        (
            parts.scheme,
            parts.netloc,
            parts.path,
            urlencode(query_pairs, doseq=True),
            parts.fragment,
        )
    )


def build_category_url(category_code: str) -> str:
    """
    여성 필터 gf=F를 항상 포함합니다.
    """
    return ensure_gender_filter_url(f"{BASE_BRAND_URL}?categoryCode={category_code}")


def guess_product_name(lines: Iterable[str], product_id: str = "") -> str:
    """
    상품 카드 안의 여러 줄 텍스트 중 상품명으로 보이는 줄을 고릅니다.
    화면 구조가 바뀌면 이 함수부터 수정하면 됩니다.
    """
    blocked_words = [
        "무신사", "MUSINSA", "STANDARD", "스탠다드",
        "원", "%", "쿠폰", "리뷰", "평점", "좋아요", "SALE",
        "무료배송", "배송", "품절",
        "멤버스데이", "단독", "옵션", "공용", "도착보장",
        "영상정보처리기기", "관리방침", "이용약관", "오프라인 전용",
    ]

    candidates: List[str] = []
    for line in lines:
        line = clean_text(line)
        if not line:
            continue
        if product_id and product_id in line:
            continue
        if any(word.lower() in line.lower() for word in blocked_words):
            continue
        if len(line) < 3:
            continue
        candidates.append(line)

    if candidates:
        # 보통 상품명은 너무 짧지 않고, 너무 길지도 않습니다.
        candidates = sorted(candidates, key=lambda x: (abs(len(x) - 24), len(x)))
        return candidates[0]

    # 마지막 안전장치
    for line in lines:
        line = clean_text(line)
        if line and "원" not in line and "%" not in line:
            return line

    return ""


def normalize_size_token(token: str) -> str:
    token = clean_text(token).upper()
    token = token.replace("ONE SIZE", "ONE SIZE")
    return token


def extract_size_from_text(text: str) -> str:
    """
    옵션 텍스트에서 사이즈처럼 보이는 값을 하나 추출합니다.
    가격/리뷰/쿠폰 같은 본문 노이즈는 거부하고, 숫자 사이즈는 화이트리스트만 통과시킵니다.
    """
    text = clean_text(text)
    if not text:
        return ""

    if any(noise in text for noise in SIZE_NOISE_WORDS):
        return ""

    for match in SIZE_WORD_RE.finditer(text):
        token = match.group(1)
        token_upper = normalize_size_token(token)

        if token_upper.isdigit():
            if token_upper in NUMERIC_SIZE_WHITELIST:
                return token_upper
            continue

        return token_upper

    return ""


def is_option_noise_text(text: str) -> bool:
    """
    옵션처럼 보이는 DOM 후보 중 리뷰/Q&A/착용 예시 텍스트를 제외합니다.
    """
    text = clean_text(text)
    if not text:
        return True
    if any(word in text for word in OPTION_NOISE_WORDS):
        return True
    if OPTION_DATE_RE.search(text):
        return True
    return False


def find_color_words(text: str) -> List[str]:
    """
    옵션 텍스트에서 색상 표준명을 추출합니다.
    한·영 표기(예: BLACK / 블랙)는 같은 색으로 묶어 한 번만 반환합니다.
    같은 위치를 짧은 alias가 다시 잡지 않도록 긴 alias부터 매칭합니다.
    """
    text_clean = clean_text(text)
    if not text_clean:
        return []

    flat: List[Tuple[str, str]] = []
    for canonical, aliases in COLOR_ALIASES.items():
        for alias in aliases:
            flat.append((alias, canonical))
    flat.sort(key=lambda pair: -len(pair[0]))

    matched_spans: List[Tuple[int, int]] = []
    found: List[str] = []
    seen: set = set()

    for alias, canonical in flat:
        for match in re.finditer(re.escape(alias), text_clean, re.IGNORECASE):
            start, end = match.start(), match.end()
            if any(s <= start < e or s < end <= e for s, e in matched_spans):
                continue
            matched_spans.append((start, end))
            if canonical not in seen:
                found.append(canonical)
                seen.add(canonical)

    return found


def extract_color_from_text(text: str) -> str:
    """
    한 옵션 라벨 안에서 감지된 색상들을 SKU의 색상 축 값 하나로 정리합니다.
    예: '01.화이트/블랙 · XL'은 두 개 SKU가 아니라 '화이트/블랙' 색상 옵션 하나입니다.
    """
    colors = find_color_words(text)
    if not colors:
        return ""
    return "/".join(colors)


def option_status_from_text(text: str, disabled: bool = False) -> str:
    text = clean_text(text).lower()
    if disabled:
        return "품절"
    if "품절" in text or "sold out" in text or "soldout" in text:
        return "품절"
    if "재입고" in text and "알림" in text:
        return "품절"
    return "판매중"


def make_sku_key(product_id: str, color: str, size: str) -> str:
    def normalize_part(value: str) -> str:
        value = clean_text(value)
        value = value.replace("/", "-")
        value = re.sub(r"\s+", "_", value)
        value = re.sub(r"[^0-9A-Za-z가-힣_\-]", "", value)
        return value or "UNKNOWN"

    return f"{normalize_part(product_id)}_{normalize_part(color)}_{normalize_part(size)}"


# =========================
# 5. 브라우저 관련 함수
# =========================

def create_driver(headless: bool) -> Any:
    """
    Selenium Firefox 드라이버를 생성합니다.
    현재 작업 환경에는 Firefox/Geckodriver가 있으므로 Firefox를 기본으로 사용합니다.
    """
    options = FirefoxOptions()
    if headless:
        options.add_argument("--headless")
    options.set_preference("intl.accept_languages", "ko-KR,ko,en-US,en")
    options.set_preference(
        "general.useragent.override",
        (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:124.0) "
            "Gecko/20100101 Firefox/124.0"
        ),
    )

    snap_firefox = Path("/snap/firefox/current/usr/lib/firefox/firefox")
    snap_geckodriver = Path("/snap/firefox/current/usr/lib/firefox/geckodriver")
    if snap_firefox.exists():
        options.binary_location = str(snap_firefox)

    geckodriver_path = str(snap_geckodriver) if snap_geckodriver.exists() else shutil.which("geckodriver")
    service = FirefoxService(executable_path=geckodriver_path) if geckodriver_path else None
    driver = webdriver.Firefox(service=service, options=options)
    driver.set_window_size(1440, 1200)
    driver.set_page_load_timeout(45)
    return driver


def safe_goto(driver: Any, url: str, timeout_seconds: int = 45) -> bool:
    url = ensure_gender_filter_url(url)
    try:
        driver.get(url)
        WebDriverWait(driver, timeout_seconds).until(
            lambda d: d.execute_script("return document.readyState") in ("interactive", "complete")
        )
        time.sleep(1.0)
        return True
    except TimeoutException:
        print(f"[경고] 페이지 로딩 시간 초과: {url}")
        return False
    except WebDriverException as exc:
        print(f"[경고] 페이지 이동 실패: {url} / {exc}")
        return False


def close_common_popups(driver: Any) -> None:
    """
    자주 보이는 팝업을 닫으려는 보조 함수입니다.
    실패해도 크롤링을 계속합니다.
    """
    xpaths = [
        "//button[contains(normalize-space(.), '닫기')]",
        "//button[contains(normalize-space(.), '확인')]",
        "//button[contains(normalize-space(.), '오늘 하루 보지 않기')]",
        "//*[@aria-label='닫기']",
        "//*[@aria-label='close']",
    ]

    for xpath in xpaths:
        try:
            for element in driver.find_elements(By.XPATH, xpath)[:3]:
                if element.is_displayed() and element.is_enabled():
                    element.click()
                    time.sleep(0.3)
                    break
        except Exception:
            pass


def parse_product_items_from_html(html: str) -> List[Dict[str, Any]]:
    """
    현재 DOM 스냅샷에서 상품 카드 후보를 BeautifulSoup으로 추출합니다.
    스크롤 중 여러 번 호출해 누적하면 가상 스크롤 누락을 줄일 수 있습니다.
    """
    soup = BeautifulSoup(html, "html.parser")
    results: List[Dict[str, Any]] = []
    seen = set()

    def is_product_href(href: str) -> bool:
        return any(pattern.search(href or "") for pattern in PRODUCT_ID_PATTERNS)

    def card_root(anchor: Any) -> Any:
        """
        한 상품 카드의 경계는 "여러 product_id를 동시에 품지 않는 가장 큰 조상".
        같은 카드 안에 이미지 링크/제목 링크/태그 칩 등이 모두 같은 상품을 가리킬 수 있으니
        anchor 개수가 아니라 _고유 product_id 수_로 카드 경계를 잡습니다.
        """
        last_good = anchor.parent or anchor
        for parent in anchor.parents:
            if not getattr(parent, "name", None):
                continue
            try:
                ids = set()
                for a in parent.find_all("a", href=True):
                    pid = extract_product_id(a.get("href", ""))
                    if pid:
                        ids.add(pid)
                unique_ids = len(ids)
            except Exception:
                unique_ids = 1
            if unique_ids <= 1:
                last_good = parent
                if parent.name in ("body", "html"):
                    break
            else:
                break
        return last_good

    def extract_card_name(root: Any, anchor: Any) -> str:
        """
        무신사 새 DOM 기준 상품명은 다음 위치 중 하나에서 정확히 노출됩니다.
        1) 카드 안 <img alt="...">  ← 이미지 alt 가 곧 상품명
        2) 텍스트 링크 <a aria-label="... 상품상세로 이동"> 의 prefix
        3) <span data-mds="Typography"> 의 텍스트
        이 순서대로 시도하고, 마지막 안전망으로 anchor 자체의 aria-label 을 사용.
        """
        if root is not None:
            # (1) 이미지 alt
            for img in root.find_all("img"):
                alt = clean_text(img.get("alt", ""))
                if alt and len(alt) >= 3 and not alt.startswith("좋아요"):
                    return alt
            # (2) 텍스트 링크의 aria-label
            for a in root.find_all("a"):
                aria = clean_text(a.get("aria-label", ""))
                if aria and "상품상세로 이동" in aria:
                    return clean_text(re.sub(r"\s*상품상세로\s*이동\s*$", "", aria))
            # (3) 무신사 디자인 시스템 Typography span
            span = root.select_one('span[data-mds="Typography"]')
            if span:
                txt = clean_text(span.get_text(" ", strip=True))
                if txt and len(txt) >= 3:
                    return txt
        # (4) 안전망: anchor 의 aria-label
        aria = clean_text(anchor.get("aria-label", "") if anchor else "")
        if aria and aria != "상품 상세로 이동":
            return clean_text(re.sub(r"\s*상품상세로\s*이동\s*$", "", aria))
        return ""

    def image_url(root: Any) -> str:
        img = root.find("img") if root else None
        if not img:
            return ""
        srcset = clean_text(img.get("srcset", ""))
        if srcset:
            return clean_text(srcset.split(",")[0].split()[0])
        for attr in ("src", "data-src", "data-original"):
            value = clean_text(img.get(attr, ""))
            if value:
                return value
        return ""

    for anchor in soup.find_all("a", href=True):
        href = ensure_gender_filter_url(anchor.get("href", ""))
        if href.startswith("/"):
            href = ensure_gender_filter_url(f"https://www.musinsa.com{href}")
        if not is_product_href(href):
            continue
        if href in seen:
            continue
        seen.add(href)

        root = card_root(anchor)
        text = root.get_text("\n", strip=True) if root else anchor.get_text("\n", strip=True)
        lines = [clean_text(line) for line in text.split("\n") if clean_text(line)]
        results.append(
            {
                "product_url": href,
                "product_name": extract_card_name(root, anchor),
                "lines": lines,
                "card_text": " | ".join(lines),
                "image_url": image_url(root),
            }
        )

    return results


def scroll_and_collect_products(
    driver: Any,
    max_scrolls: int,
    wait_ms: int = 1000,
    max_products: Optional[int] = None,
) -> List[Dict[str, Any]]:
    """
    Selenium으로 스크롤하며 각 DOM 스냅샷의 상품을 누적합니다.
    최종 DOM 한 번만 읽으면 가상 스크롤 페이지에서 상품이 빠질 수 있습니다.
    """
    items_by_key: Dict[str, Dict[str, Any]] = {}
    stable_count = 0
    last_height = 0
    last_seen_count = 0

    scroll_bar = tqdm(
        range(max_scrolls + 1),
        desc="목록 스크롤/누적",
        unit="회",
        leave=False,
        dynamic_ncols=True,
    )

    for index in scroll_bar:
        for item in parse_product_items_from_html(driver.page_source):
            product_url = ensure_gender_filter_url(item.get("product_url", ""))
            product_id = extract_product_id(product_url)
            key = product_id or product_url
            if key and key not in items_by_key:
                items_by_key[key] = item

        scroll_bar.set_postfix(상품=len(items_by_key))
        if max_products is not None and len(items_by_key) >= max_products:
            break

        try:
            height = int(driver.execute_script("return document.body.scrollHeight || 0"))
            y = int(driver.execute_script("return window.scrollY || 0"))
            viewport = int(driver.execute_script("return window.innerHeight || 900"))
            driver.execute_script("window.scrollTo(0, arguments[0]);", y + max(int(viewport * 0.9), 900))
        except WebDriverException:
            break

        time.sleep(wait_ms / 1000)

        if index == 0:
            last_height = height
            last_seen_count = len(items_by_key)
            continue

        no_new_products = len(items_by_key) == last_seen_count
        same_height = height == last_height
        stable_count = stable_count + 1 if no_new_products and same_height else 0
        if stable_count >= 5:
            break
        last_height = height
        last_seen_count = len(items_by_key)

    return list(items_by_key.values())


# =========================
# 5b. 카테고리 트리 동적 발견
# =========================

MUSINSA_CATEGORY_INDEX_URL = "https://www.musinsa.com/category/{code}"


def _extract_category_id_pairs(html: str) -> List[Tuple[str, str, str]]:
    """
    페이지에서 [data-category-id] 가 붙은 모든 노드를 긁어 (code, full_name, section) 트리플 반환.
    full_name 은 무신사가 쓰는 "대분류명|중분류명" 포맷.
    section 은 'menu_tab' (대분류 탭) 또는 'category_depth' (서브 탭).
    """
    soup = BeautifulSoup(html, "html.parser")
    out: List[Tuple[str, str, str]] = []
    seen = set()
    for el in soup.select("[data-category-id]"):
        cid = clean_text(el.get("data-category-id", ""))
        if not cid or not cid.isdigit():
            continue
        if len(cid) not in (3, 6):
            continue
        cname = clean_text(el.get("data-category-name", ""))
        section = clean_text(el.get("data-section-name", ""))
        key = (cid, cname, section)
        if key in seen:
            continue
        seen.add(key)
        out.append(key)
    return out


def _split_major_sub_name(full_name: str) -> Tuple[str, str]:
    """
    'data-category-name' 값을 (대분류명, 중분류명) 으로 분해.
    "(not set)" 이나 빈 값은 ('', '') 로 반환합니다.
    """
    name = clean_text(full_name)
    if not name or name == "(not set)":
        return "", ""
    parts = [p.strip() for p in name.split("|") if p.strip()]
    if len(parts) >= 2:
        return parts[0], parts[1]
    return "", parts[0] if parts else ""


def discover_category_tree(driver: Any) -> List[Dict[str, Any]]:
    """
    무신사 카테고리 네비게이션을 그때그때 긁어
    [{major_code, major_name, subs: [{code, name}, ...]}, ...] 트리를 만듭니다.

    1) 브랜드 페이지(예: ?categoryCode=001&gf=F)에서 'menu_tab' 영역의 3자리 코드를
       모두 수집해 브랜드가 보유한 대분류 코드 집합을 확정.
    2) 각 대분류는 무신사 공통 카테고리 페이지 https://www.musinsa.com/category/{code}?gf=F
       에서 'category_depth' 영역의 6자리 서브 코드 + 한글명을 그때그때 긁어옴.
       (브랜드 페이지의 서브 탭은 swiper로 가려지지만, 공통 카테고리 페이지는 풀 리스트를 노출.)
    """
    print("\n[카테고리 트리 발견] 대분류 코드 수집")

    seed_url = ensure_gender_filter_url(f"{BASE_BRAND_URL}?categoryCode=001")
    safe_goto(driver, seed_url)
    close_common_popups(driver)
    time.sleep(2.0)

    major_codes: List[str] = []
    seen_majors: set = set()
    for cid, _name, section in _extract_category_id_pairs(driver.page_source):
        if section == "menu_tab" and len(cid) == 3 and cid not in seen_majors:
            seen_majors.add(cid)
            major_codes.append(cid)

    if not major_codes:
        print("[카테고리 트리] ⚠ menu_tab 발견 실패 → 시드 폴백 사용")
        major_codes = list(MAJOR_CODE_SEEDS)

    print(f"[카테고리 트리] 대분류 후보: {', '.join(major_codes)}")

    tree: List[Dict[str, Any]] = []
    for major_code in major_codes:
        index_url = ensure_gender_filter_url(MUSINSA_CATEGORY_INDEX_URL.format(code=major_code))
        if not safe_goto(driver, index_url):
            continue
        close_common_popups(driver)
        time.sleep(1.5)

        major_name = ""
        subs: Dict[str, str] = {}
        for cid, full_name, _section in _extract_category_id_pairs(driver.page_source):
            mname, sname = _split_major_sub_name(full_name)
            if cid == major_code and mname and not major_name:
                major_name = mname
            if len(cid) == 6 and cid.startswith(major_code) and sname:
                if cid not in subs:
                    subs[cid] = sname
                    if mname and not major_name:
                        major_name = mname

        if not subs:
            print(f"  - {major_code}: 서브 0개 (스킵)")
            continue

        major_name = major_name or MAJOR_NAME_FALLBACK.get(major_code, major_code)
        subs_sorted = [{"code": c, "name": n} for c, n in sorted(subs.items())]
        tree.append(
            {
                "major_code": major_code,
                "major_name": major_name,
                "subs": subs_sorted,
            }
        )
        print(f"  - {major_code} {major_name}: 서브 {len(subs_sorted)}개")

    if tree:
        print(f"[카테고리 트리] 대분류 {len(tree)}개, 서브 {sum(len(m['subs']) for m in tree)}개 확정")
    else:
        print("[카테고리 트리] ⚠ 서브카테고리를 한 개도 발견하지 못했습니다.")

    return tree


def category_tree_to_master_rows(tree: List[Dict[str, Any]]) -> List[Dict[str, str]]:
    """
    동적으로 발견된 트리를 category_master 시트용 평탄한 행 리스트로 변환.
    """
    rows: List[Dict[str, str]] = []
    for major in tree:
        for sub in major["subs"]:
            rows.append(
                {
                    "major_category": major["major_name"],
                    "major_code": major["major_code"],
                    "category_code": sub["code"],
                    "raw_category": sub["name"],
                }
            )
    return rows


# =========================
# 6. 상품 목록 수집
# =========================

def collect_products_from_category(
    driver: Any,
    major_category: str,
    category_code: str,
    max_products: Optional[int],
    max_scrolls: int,
    raw_category: str = "미확인",
) -> List[ProductRow]:
    category_url = build_category_url(category_code)
    print(f"\n[목록 수집] {major_category} / {raw_category} / categoryCode={category_code}")
    print(f"  URL: {category_url}")

    if not safe_goto(driver, category_url):
        return []

    close_common_popups(driver)
    raw_items = scroll_and_collect_products(
        driver=driver,
        max_scrolls=max_scrolls,
        max_products=max_products,
    )

    rows: List[ProductRow] = []
    seen_ids_or_urls = set()
    crawl_date = now_kst_str()

    for item in tqdm(
        raw_items,
        desc="상품 카드 처리",
        unit="개",
        leave=False,
        dynamic_ncols=True,
    ):
        product_url = ensure_gender_filter_url(item.get("product_url", ""))
        product_id = extract_product_id(product_url)
        key = product_id or product_url
        if not key or key in seen_ids_or_urls:
            continue
        seen_ids_or_urls.add(key)

        lines = item.get("lines", []) or []
        card_text = clean_text(item.get("card_text", ""))
        product_name = clean_text(item.get("product_name", "")) or guess_product_name(
            lines, product_id=product_id
        )
        price = parse_int_price(card_text)
        discount_rate = parse_discount(card_text)
        image_url = clean_text(item.get("image_url", ""))

        is_soldout = "품절" if "품절" in card_text else "판매중"

        rows.append(
            ProductRow(
                crawl_date=crawl_date,
                gender_filter=GENDER_FILTER,
                brand=BRAND,
                major_category=major_category,
                raw_category=raw_category,
                category_code=category_code,
                product_id=product_id,
                product_name=product_name,
                product_url=ensure_gender_filter_url(product_url),
                price=price,
                discount_rate=discount_rate,
                image_url=image_url,
                is_soldout=is_soldout,
                source_category_url=ensure_gender_filter_url(category_url),
            )
        )

        if max_products is not None and len(rows) >= max_products:
            break

    print(f"  누적 상품 후보 수: {len(raw_items)}")
    print(f"  수집 상품 수: {len(rows)}")
    return rows


# =========================
# 7. 상세 페이지 / 옵션 수집
# =========================

def get_page_body_text(driver: Any) -> str:
    try:
        soup = BeautifulSoup(driver.page_source, "html.parser")
        body = soup.find("body")
        return clean_text(body.get_text(" ", strip=True) if body else soup.get_text(" ", strip=True))
    except Exception:
        return ""


def get_detail_title(driver: Any, fallback: str) -> str:
    """
    상세 페이지에서 상품명을 다시 확인합니다.
    """
    try:
        soup = BeautifulSoup(driver.page_source, "html.parser")
        meta_title = soup.select_one('meta[property="og:title"]') or soup.select_one('meta[name="title"]')
        title = clean_text(meta_title.get("content", "") if meta_title else "")
        if not title:
            h1 = soup.find("h1")
            title = clean_text(h1.get_text(" ", strip=True) if h1 else "")
        if not title:
            title = clean_text(driver.title)
    except Exception:
        title = ""

    title = re.sub(r"\s*\|\s*무신사.*$", "", title).strip()
    title = re.sub(r"\s*-\s*무신사.*$", "", title).strip()
    return title or fallback


def get_breadcrumb_like_text(driver: Any) -> str:
    """
    카테고리명은 본문 전체보다 빵부스러기 메뉴, 즉 경로 영역에서 찾는 것이 더 안전합니다.
    """
    try:
        soup = BeautifulSoup(driver.page_source, "html.parser")
        chunks: List[str] = []
        selectors = [
            "nav",
            '[class*="breadcrumb"]',
            '[class*="Breadcrumb"]',
            '[class*="category"]',
            '[class*="Category"]',
        ]
        for selector in selectors:
            for element in soup.select(selector):
                text = clean_text(element.get_text(" ", strip=True))
                if text and len(text) < 1000:
                    chunks.append(text)
        return clean_text("\n".join(chunks))
    except Exception:
        return ""


def collect_option_text_candidates(driver: Any) -> List[Dict[str, Any]]:
    """
    상세 페이지에서 옵션처럼 보이는 텍스트를 모읍니다.
    너무 완벽하게 잡으려 하지 않고, 데모에서는 후보를 넓게 잡은 뒤 파이썬에서 걸러냅니다.
    """
    try:
        soup = BeautifulSoup(driver.page_source, "html.parser")
    except Exception:
        return []

    selectors = [
        "button",
        "option",
        "li",
        '[role="option"]',
        '[class*="option"]',
        '[class*="Option"]',
        '[data-testid*="option"]',
        '[data-test*="option"]',
    ]
    nodes = []
    for selector in selectors:
        try:
            nodes.extend(soup.select(selector))
        except Exception:
            pass

    cleaned: List[Dict[str, Any]] = []
    seen = set()
    for node in nodes:
        text = clean_text(node.get_text(" ", strip=True))
        if not text or len(text) > 80:
            continue
        if is_option_noise_text(text):
            continue

        outer_html = str(node)
        disabled = (
            node.has_attr("disabled")
            or node.get("aria-disabled") == "true"
            or bool(re.search(r"disabled|soldout|sold-out|품절", outer_html, re.IGNORECASE))
        )

        context_parts: List[str] = []
        cur = node
        for _ in range(4):
            if not cur:
                break
            context_parts.extend(cur.get("class", []))
            context_parts.append(clean_text(cur.get("id", "")))
            context_parts.append(clean_text(cur.get("aria-label", "")))
            context_parts.append(clean_text(cur.get("data-testid", "")))
            context_parts.append(clean_text(cur.get("data-test", "")))
            cur = cur.parent
        context_text = clean_text(" ".join(context_parts))
        is_purchase_option = bool(
            re.search(r"구매옵션|선택|옵션|option|Option|select|Select", f"{text} {context_text}")
            and not re.search(r"상품상세문의|답변완료|리뷰|후기|착용", text)
        )

        key = (text, disabled, is_purchase_option)
        if key in seen:
            continue
        seen.add(key)
        cleaned.append(
            {
                "text": text,
                "disabled": disabled,
                "is_purchase_option": is_purchase_option,
            }
        )

    return cleaned


def parse_options_from_candidates(
    product_id: str,
    product_name: str,
    major_category: str,
    raw_category: str,
    product_url: str,
    candidates: List[Dict[str, Any]],
) -> Tuple[List[OptionRow], List[SkuRow]]:
    """
    옵션 후보 텍스트에서 color, size를 추정합니다.

    이 데모의 핵심 원칙:
    - 원본 옵션명 option_name은 최대한 그대로 보관합니다.
    - 실제 색상-사이즈 조합을 확인한 후보만 SKU로 계산합니다.
    - 조합을 확정하지 못한 후보는 options_raw 검증용으로만 남깁니다.
    """
    option_rows: List[OptionRow] = []
    sku_rows: List[SkuRow] = []

    normalized_url = ensure_gender_filter_url(product_url)
    ordered_candidates = sorted(
        candidates,
        key=lambda item: 0 if item.get("is_purchase_option", False) else 1,
    )
    purchase_candidates = [
        item for item in ordered_candidates
        if item.get("is_purchase_option", False)
    ]
    candidate_groups = [purchase_candidates, ordered_candidates] if purchase_candidates else [ordered_candidates]

    # 구매 옵션 영역에서 나온 색상-사이즈 조합을 우선 사용합니다.
    for group in candidate_groups:
        combo_seen = set()
        group_option_rows: List[OptionRow] = []

        for item in group:
            text = clean_text(item.get("text", ""))
            disabled = bool(item.get("disabled", False))

            if is_option_noise_text(text):
                continue

            size = extract_size_from_text(text)
            color = extract_color_from_text(text)
            status = option_status_from_text(text, disabled=disabled)

            if not (size and color):
                continue

            key = (color, size)
            if key in combo_seen:
                continue
            combo_seen.add(key)

            group_option_rows.append(
                OptionRow(
                    product_id=product_id,
                    product_name=product_name,
                    major_category=major_category,
                    raw_category=raw_category,
                    color=color,
                    size=size,
                    option_name=text,
                    option_status=status,
                    option_price=parse_int_price(text),
                    stock_qty=None,
                    product_url=normalized_url,
                )
            )

        if group_option_rows:
            option_rows = group_option_rows
            break

    if option_rows:
        for opt in option_rows:
            sku_rows.append(
                SkuRow(
                    major_category=major_category,
                    raw_category=raw_category,
                    product_id=product_id,
                    product_name=product_name,
                    color=opt.color,
                    size=opt.size,
                    sku_key=make_sku_key(product_id, opt.color, opt.size),
                    sku_status=opt.option_status,
                    product_url=normalized_url,
                )
            )
        return option_rows, sku_rows

    # 실제 조합을 확정하지 못한 경우에는 SKU를 만들지 않고 검증용 후보만 남깁니다.
    diagnostic_seen = set()
    for item in ordered_candidates:
        text = clean_text(item.get("text", ""))
        if is_option_noise_text(text):
            continue

        color = extract_color_from_text(text) or "색상 미확인"
        size = extract_size_from_text(text)
        if not size and color == "색상 미확인":
            continue
        size = size or "사이즈 미확인"

        key = (color, size, text)
        if key in diagnostic_seen:
            continue
        diagnostic_seen.add(key)

        option_rows.append(
            OptionRow(
                product_id=product_id,
                product_name=product_name,
                major_category=major_category,
                raw_category=raw_category,
                color=color,
                size=size,
                option_name=text,
                option_status="옵션 후보",
                option_price=parse_int_price(text),
                stock_qty=None,
                product_url=normalized_url,
            )
        )

    if not option_rows:
        option_rows.append(
            OptionRow(
                product_id=product_id,
                product_name=product_name,
                major_category=major_category,
                raw_category=raw_category,
                color="색상 미확인",
                size="사이즈 미확인",
                option_name="옵션 미확인",
                option_status="옵션 미확인",
                option_price=None,
                stock_qty=None,
                product_url=normalized_url,
            )
        )

    return option_rows, sku_rows


def collect_detail_and_options(
    driver: Any,
    product: ProductRow,
) -> Tuple[ProductRow, List[OptionRow], List[SkuRow]]:
    if not product.product_url:
        return product, [], []

    product_url = ensure_gender_filter_url(product.product_url)
    source_category_url = ensure_gender_filter_url(product.source_category_url)

    if not safe_goto(driver, product_url):
        return product, [], []

    close_common_popups(driver)

    product_name = get_detail_title(driver, fallback=product.product_name)
    body_text = get_page_body_text(driver)
    # 서브카테고리 단위로 크롤하므로 raw_category는 목록 수집 시점에 이미 정확히 라벨링됨.
    raw_category = product.raw_category or "미확인"

    # 목록에서 가격을 못 가져왔으면 상세 본문에서 한 번 더 시도
    price = product.price or parse_int_price(body_text)

    updated_product = ProductRow(
        crawl_date=product.crawl_date,
        gender_filter=product.gender_filter,
        brand=product.brand,
        major_category=product.major_category,
        raw_category=raw_category,
        category_code=product.category_code,
        product_id=product.product_id,
        product_name=product_name,
        product_url=product_url,
        price=price,
        discount_rate=product.discount_rate,
        image_url=product.image_url,
        is_soldout=product.is_soldout,
        source_category_url=source_category_url,
    )

    candidates = collect_option_text_candidates(driver)
    option_rows, sku_rows = parse_options_from_candidates(
        product_id=updated_product.product_id,
        product_name=updated_product.product_name,
        major_category=updated_product.major_category,
        raw_category=updated_product.raw_category,
        product_url=updated_product.product_url,
        candidates=candidates,
    )

    return updated_product, option_rows, sku_rows


# =========================
# 8. 엑셀 저장
# =========================

def dataclass_list_to_dicts(rows: Iterable[Any]) -> List[Dict[str, Any]]:
    return [row.__dict__.copy() for row in rows]


def make_model_key_from_values(
    major_category: Any,
    raw_category: Any,
    product_id: Any,
    product_url: Any,
) -> str:
    major_value = "" if pd.isna(major_category) else major_category
    raw_value = "" if pd.isna(raw_category) else raw_category
    product_id_value = "" if pd.isna(product_id) else product_id
    product_url_value = "" if pd.isna(product_url) else product_url
    product_identity = clean_text(product_id_value) or ensure_gender_filter_url(product_url_value)
    return "|".join(
        [
            clean_text(major_value),
            clean_text(raw_value),
            product_identity,
        ]
    )


def add_model_key_column(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        df["_model_key"] = []
        return df

    df = df.copy()
    df["_model_key"] = df.apply(
        lambda row: make_model_key_from_values(
            row.get("major_category", ""),
            row.get("raw_category", ""),
            row.get("product_id", ""),
            row.get("product_url", ""),
        ),
        axis=1,
    )
    return df


def normalize_url_columns(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    for column in ("product_url", "source_category_url"):
        if column in df.columns:
            df[column] = df[column].map(
                lambda value: "" if pd.isna(value) else ensure_gender_filter_url(value)
            )
    return df


def clean_model_name(raw_name: str) -> str:
    """
    상품 카드/상세에서 가져온 긴 상품명을 보고서 표시용 모델명으로 정리합니다.
    브랜드 prefix와 " - 사이즈 & 후기" 같은 보조 표기를 제거합니다.
    """
    if not raw_name:
        return ""
    name = clean_text(raw_name)
    name = re.sub(r"^무신사\s*스탠다드\s*\(MUSINSA\s*STANDARD\)\s*", "", name, flags=re.IGNORECASE)
    name = re.sub(r"\s*-\s*사이즈\s*&?\s*후기.*$", "", name)
    name = re.sub(r"\s*-\s*후기.*$", "", name)
    return name.strip()


def format_price_summary(prices: Iterable[Optional[int]]) -> str:
    """
    중분류 단위 가격을 '최고가 / 최저가 / 평균가원' 형식으로 반환합니다.
    유효한 가격이 없으면 '-' 를 반환합니다.
    """
    valid = [int(p) for p in prices if p is not None]
    if not valid:
        return "-"
    return f"{max(valid):,} / {min(valid):,} / {int(round(sum(valid) / len(valid))):,}원"


def build_model_summary_df(
    products: List[ProductRow],
    sku_rows: List[SkuRow],
) -> pd.DataFrame:
    """
    상품(모델) 단위 요약. SKU_요약 시트와 model_summary 시트의 공통 입력입니다.
    """
    columns = [
        "major_category", "raw_category", "product_id", "product_name",
        "color_count", "size_count", "sku_count",
        "available_sku_count", "soldout_sku_count",
        "price", "product_url",
    ]

    if not products:
        return pd.DataFrame(columns=columns)

    prod_df = normalize_url_columns(pd.DataFrame(dataclass_list_to_dicts(products)))
    sku_df = normalize_url_columns(pd.DataFrame(dataclass_list_to_dicts(sku_rows)))
    prod_df = add_model_key_column(prod_df)

    if sku_df.empty:
        prod_df = prod_df.assign(
            color_count=0, size_count=0, sku_count=0,
            available_sku_count=0, soldout_sku_count=0,
        )
    else:
        sku_df = add_model_key_column(sku_df)
        agg = sku_df.groupby("_model_key").agg(
            color_count=("color", "nunique"),
            size_count=("size", "nunique"),
            sku_count=("_model_key", "count"),
            available_sku_count=("sku_status", lambda s: int((s == "판매중").sum())),
            soldout_sku_count=("sku_status", lambda s: int((s == "품절").sum())),
        ).reset_index()
        prod_df = prod_df.merge(agg, on="_model_key", how="left")
        for col in ("color_count", "size_count", "sku_count", "available_sku_count", "soldout_sku_count"):
            prod_df[col] = prod_df[col].fillna(0).astype(int)

    return prod_df.reindex(columns=columns)


def build_sku_summary_sheet(
    ws: Any,
    products: List[ProductRow],
    sku_rows: List[SkuRow],
    category_tree: Optional[List[Dict[str, Any]]] = None,
) -> None:
    """
    style.xlsx 양식의 보고서형 SKU_요약 시트를 채웁니다.
    A열(대분류) / B,F,G,H,I열(중분류) 단위로 세로 병합합니다.
    """
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    headers = [
        "카테고리", "중분류", "모델명", "컬러 수", "사이즈 수",
        "SKU 개수", "SKU 비중", "최고/최저/평균가", "베이직/뉴베이직/트렌디",
    ]
    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="D9D9D9")
    header_font = Font(bold=True, size=11)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)

    for col_idx, label in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.alignment = center
        cell.fill = header_fill
        cell.border = border
    ws.row_dimensions[1].height = 30
    widths = {1: 18, 2: 16, 3: 38, 4: 8, 5: 8, 6: 12, 7: 12, 8: 26, 9: 22}
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"

    summary_df = build_model_summary_df(products, sku_rows)
    if summary_df.empty:
        return

    total_sku = int(summary_df["sku_count"].sum())

    tree = category_tree or []
    major_order: List[str] = [m["major_name"] for m in tree]
    raw_order_by_major: Dict[str, List[str]] = {
        m["major_name"]: [s["name"] for s in m["subs"]] for m in tree
    }

    # 트리에 없지만 데이터에는 있는 대분류도 뒤에 붙임
    for extra in summary_df["major_category"].dropna().unique().tolist():
        if extra and extra not in major_order:
            major_order.append(extra)
            raw_order_by_major.setdefault(extra, [])

    current_row = 2
    for major_cat in major_order:
        major_df = summary_df[summary_df["major_category"] == major_cat]

        major_total = int(major_df["sku_count"].sum())
        major_pct = major_total / total_sku if total_sku else 0

        raw_master = raw_order_by_major.get(major_cat, [])
        raw_in_data = list(major_df["raw_category"].unique())
        raw_sequence = (
            raw_master
            + [r for r in raw_in_data if r not in raw_master]
        )
        if not raw_sequence:
            continue

        major_start = current_row

        for raw_cat in raw_sequence:
            raw_df = major_df[major_df["raw_category"] == raw_cat].sort_values("product_id")

            raw_total = int(raw_df["sku_count"].sum())
            raw_pct = raw_total / total_sku if total_sku else 0
            price_summary = format_price_summary(raw_df["price"].tolist())
            style_summary = "미분류: 100%"

            raw_start = current_row
            if raw_df.empty:
                ws.cell(row=current_row, column=3, value="")
                ws.cell(row=current_row, column=4, value=0)
                ws.cell(row=current_row, column=5, value=0)
                current_row += 1
            else:
                for _, model in raw_df.iterrows():
                    ws.cell(row=current_row, column=3, value=clean_model_name(model["product_name"]))
                    ws.cell(row=current_row, column=4, value=int(model["color_count"]))
                    ws.cell(row=current_row, column=5, value=int(model["size_count"]))
                    current_row += 1
            raw_end = current_row - 1

            ws.cell(row=raw_start, column=2, value=raw_cat)
            ws.cell(row=raw_start, column=6, value=raw_total)
            pct_cell = ws.cell(row=raw_start, column=7, value=raw_pct)
            pct_cell.number_format = "0.00%"
            ws.cell(row=raw_start, column=8, value=price_summary)
            ws.cell(row=raw_start, column=9, value=style_summary)

            if raw_end > raw_start:
                for col in (2, 6, 7, 8, 9):
                    ws.merge_cells(
                        start_row=raw_start, end_row=raw_end,
                        start_column=col, end_column=col,
                    )

        major_end = current_row - 1
        if major_end >= major_start:
            ws.cell(
                row=major_start,
                column=1,
                value=f"{major_cat}\n(총 SKU: {major_total:,} / {major_pct:.1%})",
            )
            if major_end > major_start:
                ws.merge_cells(
                    start_row=major_start, end_row=major_end,
                    start_column=1, end_column=1,
                )

    last_row = current_row - 1
    if last_row >= 2:
        for row in ws.iter_rows(min_row=2, max_row=last_row, min_col=1, max_col=9):
            for cell in row:
                cell.border = border
                cell.alignment = left_wrap if cell.column == 3 else center


def adjust_excel_format(output_path: Path, skip_sheets: Optional[Set[str]] = None) -> None:
    """
    openpyxl로 raw 시트들에 헤더 스타일·자동필터·열 너비를 적용합니다.
    skip_sheets에 들어 있는 시트(예: SKU_요약)는 건드리지 않습니다.
    """
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        return

    skip = skip_sheets or set()
    wb = load_workbook(output_path)

    header_fill = PatternFill("solid", fgColor="EDEDED")
    header_font = Font(bold=True)

    for ws in wb.worksheets:
        if ws.title in skip:
            continue
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for column_cells in ws.columns:
            max_len = 0
            col_letter = column_cells[0].column_letter
            for cell in column_cells:
                value = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(value))
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 45)

    wb.save(output_path)


def save_excel(
    output_path: Path,
    products: List[ProductRow],
    options: List[OptionRow],
    sku_rows: List[SkuRow],
    category_tree: Optional[List[Dict[str, Any]]] = None,
) -> None:
    """
    SKU_요약(보고서형) 시트를 맨 앞에 두고, 모델/원본 시트를 뒤에 붙입니다.
    """
    from openpyxl import load_workbook

    output_path.parent.mkdir(parents=True, exist_ok=True)

    master_rows = category_tree_to_master_rows(category_tree or [])
    category_df = pd.DataFrame(
        master_rows,
        columns=["major_category", "major_code", "category_code", "raw_category"],
    )
    products_df = normalize_url_columns(pd.DataFrame(dataclass_list_to_dicts(products)))
    options_df = normalize_url_columns(pd.DataFrame(dataclass_list_to_dicts(options)))
    sku_df = normalize_url_columns(pd.DataFrame(dataclass_list_to_dicts(sku_rows)))
    model_summary_df = normalize_url_columns(build_model_summary_df(products, sku_rows))

    products_df = products_df.reindex(columns=list(ProductRow.__annotations__.keys()))
    options_df = options_df.reindex(columns=list(OptionRow.__annotations__.keys()))
    sku_df = sku_df.reindex(columns=list(SkuRow.__annotations__.keys()))

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        model_summary_df.to_excel(writer, sheet_name="model_summary", index=False)
        products_df.to_excel(writer, sheet_name="products_raw", index=False)
        options_df.to_excel(writer, sheet_name="options_raw", index=False)
        sku_df.to_excel(writer, sheet_name="sku_result", index=False)
        category_df.to_excel(writer, sheet_name="category_master", index=False)

    wb = load_workbook(output_path)
    summary_ws = wb.create_sheet("SKU_요약", 0)
    build_sku_summary_sheet(summary_ws, products, sku_rows, category_tree=category_tree)
    wb.save(output_path)

    adjust_excel_format(output_path, skip_sheets={"SKU_요약"})


# =========================
# 9. 메인 실행
# =========================

def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="무신사 스탠다드 여성 상품 SKU 계산용 엑셀 생성기 데모"
    )

    parser.add_argument(
        "--output",
        default=DEFAULT_OUTPUT,
        help=f"저장할 엑셀 파일명. 기본값: {DEFAULT_OUTPUT}",
    )
    parser.add_argument(
        "--max-products",
        type=int,
        default=10,
        help="카테고리별 최대 상품 수. 전체 수집은 0 입력. 기본값: 10",
    )
    parser.add_argument(
        "--max-scrolls",
        type=int,
        default=80,
        help="상품 목록 페이지에서 스크롤하며 누적 수집할 최대 횟수. 기본값: 80",
    )
    parser.add_argument(
        "--delay",
        type=float,
        default=0.8,
        help="상세 페이지 사이 대기 시간, 초 단위. 기본값: 0.8",
    )
    parser.add_argument(
        "--headless",
        action="store_true",
        help="브라우저 화면을 띄우지 않고 실행합니다.",
    )
    parser.add_argument(
        "--skip-options",
        action="store_true",
        help="상세 페이지 옵션 수집을 건너뛰고 상품 목록만 저장합니다.",
    )

    return parser.parse_args()


def dedupe_products(products: List[ProductRow]) -> List[ProductRow]:
    """
    같은 상품이 중복으로 잡히는 경우를 제거합니다.
    대분류가 다르면 별도 행으로 유지합니다.
    """
    result: List[ProductRow] = []
    seen = set()

    for product in products:
        key = (
            product.major_category,
            product.product_id or product.product_url,
        )
        if key in seen:
            continue
        seen.add(key)
        result.append(product)

    return result


def main() -> None:
    args = parse_args()

    max_products = None if args.max_products == 0 else args.max_products
    output_path = Path(args.output)

    print("=" * 70)
    print("무신사 스탠다드 여성 SKU 생성기 데모")
    print("=" * 70)
    print(f"브랜드: {BRAND}")
    print(f"여성 필터: gf={GENDER_FILTER} 고정")
    print(f"카테고리: 런타임에 동적 발견 (서브카테고리 단위 크롤)")
    print(f"카테고리별 최대 상품 수: {'전체' if max_products is None else max_products}")
    print(f"옵션 수집: {'건너뜀' if args.skip_options else '실행'}")
    print(f"저장 파일: {output_path}")
    print("=" * 70)

    all_products: List[ProductRow] = []
    all_options: List[OptionRow] = []
    all_skus: List[SkuRow] = []

    driver = create_driver(headless=args.headless)
    category_tree: List[Dict[str, Any]] = []

    try:
        category_tree = discover_category_tree(driver)
        if not category_tree:
            print("[경고] 카테고리 트리가 비어 있어 크롤을 중단합니다.")
            return

        # 서브카테고리 단위 작업 단위 평탄화
        work_units: List[Dict[str, str]] = []
        for major in category_tree:
            for sub in major["subs"]:
                work_units.append(
                    {
                        "major_category": major["major_name"],
                        "category_code": sub["code"],
                        "raw_category": sub["name"],
                    }
                )

        seen_product_ids: set = set()
        for unit in tqdm(
            work_units,
            desc="서브카테고리 수집",
            unit="개",
            dynamic_ncols=True,
        ):
            rows = collect_products_from_category(
                driver=driver,
                major_category=unit["major_category"],
                category_code=unit["category_code"],
                max_products=max_products,
                max_scrolls=args.max_scrolls,
                raw_category=unit["raw_category"],
            )
            # product_id 기준 전역 dedupe — 한 상품이 여러 서브에 들어가도 1번만 카운트.
            for row in rows:
                key = row.product_id or row.product_url
                if not key or key in seen_product_ids:
                    continue
                seen_product_ids.add(key)
                all_products.append(row)
            time.sleep(args.delay)

        print(f"\n[목록 수집 완료] 중복 제거 후 상품 수: {len(all_products)}")

        if not args.skip_options:
            updated_products: List[ProductRow] = []

            print("\n[상세/옵션 수집 시작]")
            for product in tqdm(
                all_products,
                desc="상세 페이지",
                unit="개",
                dynamic_ncols=True,
            ):
                updated_product, option_rows, sku_rows = collect_detail_and_options(
                    driver=driver,
                    product=product,
                )
                updated_products.append(updated_product)
                all_options.extend(option_rows)
                all_skus.extend(sku_rows)
                time.sleep(args.delay)

            all_products = updated_products

    finally:
        driver.quit()

    print("\n[엑셀 저장]")
    save_excel(
        output_path=output_path,
        products=all_products,
        options=all_options,
        sku_rows=all_skus,
        category_tree=category_tree,
    )

    print("=" * 70)
    print("완료")
    print(f"상품 행 수: {len(all_products)}")
    print(f"옵션 행 수: {len(all_options)}")
    print(f"SKU 행 수: {len(all_skus)}")
    print(f"저장 위치: {output_path.resolve()}")
    print("=" * 70)


if __name__ == "__main__":
    main()
